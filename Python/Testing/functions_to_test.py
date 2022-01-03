import os
from typing import Union

import openpyxl

os.chdir(os.path.dirname(__file__))


def addition(num1: float, num2: float) -> float:
    return num1 + num2


def divide(num1: float, num2: float) -> float:
    return num1 / num2


def divide_with_try_except(num1: float, num2: float) -> Union[float, str]:
    try:
        return num1 / num2
    except ZeroDivisionError:
        return "Can't divide by zero"


def read_text_file_contents(filename: str) -> str:
    with open(filename) as f:
        contents = f.readlines()
        return "".join(contents)


def write_data_to_text_file(filename: str, input: str) -> None:
    with open(filename, 'w') as f:
        f.write(input)
        return None


def read_excel_file_contents(filename: str, sheetname: str, cell: str) -> str:
    wb = openpyxl.load_workbook(filename, read_only=True)
    ws = wb[sheetname]
    return ws[cell].value


def load_workbook(filename):
    wb = openpyxl.load_workbook(filename, read_only=True)
    print(wb.sheetnames)
    return wb

def load_worksheet(wb):
    ws = wb